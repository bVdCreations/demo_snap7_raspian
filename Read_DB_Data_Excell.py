import openpyxl
from openpyxl.utils import get_column_letter


class ReadDB_Data:
    """
    this reads the db files in an excell and stores it in a multi level dict

    """

    def __init__(self, file_name, file_dir=''):
        """

        :param file_name: the name of the excell file (standard 'DBs_PLC_{}.xlsx'.format plcName
        :param file_dir: the directory of the excell file
        """
        self._fileName = file_name
        self._file_dire = file_dir
        self._workbook = openpyxl.load_workbook(self._file_dire+self._fileName)

    def read_data_dbs(self):
        """ returns a dict with the as key the DB names and as value:
         a dict with all the variables of the db
        """
        returndict = dict()
        for sheetname in self._workbook.sheetnames:
            if 'DB' in sheetname:
                returndict.update({sheetname: dict()})
                sheet_object = self._workbook[sheetname]
                last_entry = self.get_last_entry_sheet(sheet_object, offset=-1)

                for rowOfCellObjects in sheet_object['A3':last_entry]:
                    namerow = rowOfCellObjects[1].value
                    returndict[sheetname].update({namerow: dict()})
                    for cellObj in rowOfCellObjects:

                        if cellObj.value is not None:
                            returndict[sheetname][namerow].update(
                                {sheet_object.cell(column=cellObj.col_idx, row=1).value: cellObj.value})

        return returndict

    def read_info_plc(self):
        """ returns a dict with the keys IP_adress, rack and slot with it's values"""
        returndict = dict()
        sheetnames = self._workbook.sheetnames
        if 'info_PLC' in sheetnames:

            sheet_object = self._workbook['info_PLC']
            for rowOfCellObjects in sheet_object['A1':self.get_last_entry_sheet(sheet_object)]:
                # loop through sheet row by row
                key_info = rowOfCellObjects[0].value

                if key_info == 'IP_adress':
                    value_info = rowOfCellObjects[1].value
                else:
                    value_info = int(rowOfCellObjects[1].value)

                returndict.update({key_info: value_info})
            return returndict
        else:
            raise FileNotFoundError('"info_PLC" can not be found in the sheetnames : {}'.format(sheetnames))

    def read_i_o(self):
        """ returns a dict with the keys Inputs and Outputs
        the values of those keys are the """

        sheetnames = self._workbook.sheetnames
        if 'Symbol_table' in sheetnames:

            return_dict = {'Input': dict(), 'Output': dict()}

            # loop through the rows of the sheet
            sheet_object = self._workbook['Symbol_table']
            last_entry = self.get_last_entry_sheet(sheet_object)
            for rowOfCellObjects in sheet_object['A2':last_entry]:
                symbol_name = rowOfCellObjects[0].value
                adress_io = rowOfCellObjects[1].value

                # check if its an input or output
                if adress_io[0] == 'I':
                    store = 'Input'
                elif adress_io[0] == 'Q':
                    store = 'Output'
                else:
                    raise ValueError('the adress "{}" should start with I or Q'.format(adress_io))

                return_dict[store].update({symbol_name: dict()})

                # loop through the cells in the row
                for cellObj in rowOfCellObjects:
                    
                    if cellObj.value is not None:
                        return_dict[store][symbol_name].update(
                            {sheet_object.cell(column=cellObj.col_idx, row=1).value: cellObj.value})
            return return_dict
        else:
            raise FileNotFoundError('"Symbol_table" can not be found in the sheetnames : {}'.format(sheetnames))

    def get_last_entry_sheet(self, sheet_object: openpyxl, offset=0):
        # find the maximum range of data in the sheet
        return self.get_last_entry_column(sheet_object, row=1) + \
               str(self.get_last_entry_row(sheet_object, start_row=1)+offset)

    @staticmethod
    def get_last_entry_row(sheet_object: openpyxl, start_row=1, column=1):
        # find the row number of the last entry in the given column
        row = start_row
        last_entry_row = row
        while sheet_object.cell(column=column, row=row).value is not None:
            last_entry_row = row
            row += 1
        return last_entry_row

    @staticmethod
    def get_last_entry_column(sheet_object: openpyxl, start_column=1, row=1):
        # find the column number of the last entry the given row
        column = start_column
        last_entry_column = column
        while sheet_object.cell(column=column, row=row).value is not None:
            last_entry_column = column
            column += 1
        return get_column_letter(last_entry_column)


if '__main__' == __name__:

    for key, value in ReadDB_Data('DBs_PLC_300.xlsx', file_dir='I_O_info_plc//').read_info_plc().items():
        print('key = {}   value = {}'.format(key, value))

    for db, values in ReadDB_Data('DBs_PLC_300.xlsx', file_dir='I_O_info_plc//').read_data_dbs().items():
        print('key ={}'.format(db))

        for keyvar, value in values.items():
            print('\tkeyvar ={}'.format(keyvar))
            print('\ttype ={}'.format(value['Type']))
        print('-'*10)

    print('='*20)
    for i_or_o, values in ReadDB_Data('DBs_PLC_300.xlsx', file_dir='I_O_info_plc//').read_i_o().items():
        print('key ={}'.format(i_or_o))

        for keyvar, value in values.items():
            print('\tkeyvar ={}'.format(keyvar))
            print('\ttype ={}'.format(value['Data type']))
        print('-'*10)
